from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from fastapi.responses import JSONResponse, FileResponse
from typing import Dict, Any, Optional, List
import json
import logging
from pathlib import Path
import tempfile
import os
from datetime import datetime

from services.enhanced_document_service import EnhancedDocumentService
from models.request_models import DocumentProcessingRequest
from pydantic import BaseModel

logger = logging.getLogger(__name__)

# Router untuk document manipulation
router = APIRouter(tags=["document-manipulation"])

# Global service instance - will be injected from main app
_enhanced_document_service: Optional[EnhancedDocumentService] = None

def get_enhanced_document_service() -> EnhancedDocumentService:
    """Dependency to get enhanced document service"""
    if _enhanced_document_service is None:
        raise HTTPException(status_code=500, detail="Enhanced document service not initialized")
    return _enhanced_document_service

def set_enhanced_document_service(service: EnhancedDocumentService):
    """Set the enhanced document service instance"""
    global _enhanced_document_service
    _enhanced_document_service = service

# Request models
class DocumentManipulationRequest(BaseModel):
    file_id: str
    settings: Dict[str, Any]
    preview_only: bool = False
    
    class Config:
        schema_extra = {
            "example": {
                "file_id": "doc_123456",
                "settings": {
                    "color_mode": "black_white",
                    "convert_to_bw": True,
                    "page_range": "1-5",
                    "brightness": 10,
                    "contrast": 5,
                    "paper_size": "A4",
                    "orientation": "portrait",
                    "margin_top": 20,
                    "margin_bottom": 20,
                    "margin_left": 20,
                    "margin_right": 20,
                    "fit_to_page": True,
                    "center_horizontally": True,
                    "center_vertically": True
                },
                "preview_only": False
            }
        }

class PDFSplitRequest(BaseModel):
    file_id: str
    split_type: str = "pages"  # 'pages', 'range'
    split_value: int = 1  # pages per file
    ranges: Optional[List[str]] = None  # for range split: ["1-3", "4-6"]
    
    class Config:
        schema_extra = {
            "example": {
                "file_id": "doc_123456",
                "split_type": "pages",
                "split_value": 2,
                "ranges": ["1-3", "4-6", "7-10"]
            }
        }

class DocumentConversionRequest(BaseModel):
    file_id: str
    target_format: str = "pdf"
    conversion_settings: Optional[Dict[str, Any]] = None
    
    class Config:
        schema_extra = {
            "example": {
                "file_id": "doc_123456",
                "target_format": "pdf",
                "conversion_settings": {
                    "quality": "high",
                    "preserve_formatting": True
                }
            }
        }

class PreviewRequest(BaseModel):
    file_id: str
    page_number: int = 1
    settings: Optional[Dict[str, Any]] = None
    
    class Config:
        schema_extra = {
            "example": {
                "file_id": "doc_123456",
                "page_number": 1,
                "settings": {
                    "color_mode": "grayscale",
                    "brightness": 0,
                    "contrast": 0
                }
            }
        }

class ExcelPreviewRequest(BaseModel):
    file_id: str
    sheet_name: Optional[str] = None
    sheet_index: Optional[int] = None
    
    class Config:
        schema_extra = {
            "example": {
                "file_id": "excel_123456",
                "sheet_name": "Sheet1",
                "sheet_index": 0
            }
        }

class SpreadsheetUploadRequest(BaseModel):
    preserve_formatting: bool = True
    max_rows: int = 100
    max_columns: int = 20

class ExcelToPDFRequest(BaseModel):
    file_id: str
    preserve_formatting: bool = True
    preserve_charts: bool = True
    preserve_images: bool = True
    preserve_formulas: bool = True
    preserve_colors: bool = True
    preserve_fonts: bool = True
    preserve_borders: bool = True
    preserve_alignment: bool = True
    fit_to_page: bool = True
    high_quality: bool = True
    include_all_sheets: bool = True
    page_orientation: str = "auto"  # auto, portrait, landscape
    paper_size: str = "A4"  # A4, A3, A5, Letter, Legal
    margins: str = "normal"  # normal, narrow, wide
    conversion_method: str = "auto"  # auto, com, openpyxl, basic50
    
    class Config:
        schema_extra = {
            "example": {
                "preserve_formatting": True,
                "max_rows": 1000,
                "max_columns": 50
            }
        }

# Helper function to get file path from file_id
def get_file_path_from_id(file_id: str) -> str:
    """Get actual file path from file_id"""
    # This should integrate with your existing file service
    # For now, assuming files are stored in uploads directory
    uploads_dir = Path("uploads")
    
    # Try to find file with this ID - file_id is the actual filename with timestamp
    # First try exact match
    exact_path = uploads_dir / file_id
    if exact_path.exists() and exact_path.is_file():
        return str(exact_path)
    
    # If not exact match, try pattern matching for files ending with the file_id
    # This handles cases where file_id might be the original filename without timestamp
    for file_path in uploads_dir.glob(f"*_{file_id}"):
        if file_path.is_file():
            return str(file_path)
    
    # Also try pattern matching for files containing the file_id
    for file_path in uploads_dir.glob(f"*{file_id}*"):
        if file_path.is_file():
            return str(file_path)
    
    # If not found, try temp directory
    temp_dir = Path("temp")
    
    # Try exact match in temp
    exact_temp_path = temp_dir / file_id
    if exact_temp_path.exists() and exact_temp_path.is_file():
        return str(exact_temp_path)
    
    # Try pattern matching in temp
    for file_path in temp_dir.glob(f"*_{file_id}"):
        if file_path.is_file():
            return str(file_path)
    
    for file_path in temp_dir.glob(f"*{file_id}*"):
        if file_path.is_file():
            return str(file_path)
    
    # Try server root directory as fallback
    server_root = Path(".")
    exact_root_path = server_root / file_id
    if exact_root_path.exists() and exact_root_path.is_file():
        return str(exact_root_path)
    
    # Try pattern matching in server root
    for file_path in server_root.glob(f"*{file_id}*"):
        if file_path.is_file():
            return str(file_path)
    
    # Log available files for debugging
    logger.error(f"File with ID {file_id} not found")
    logger.error(f"Available files in uploads: {[f.name for f in uploads_dir.glob('*') if f.is_file()]}")
    logger.error(f"Available files in temp: {[f.name for f in temp_dir.glob('*') if f.is_file()]}")
    logger.error(f"Available files in server root: {[f.name for f in server_root.glob('*') if f.is_file()]}")
    
    raise HTTPException(status_code=404, detail=f"File with ID {file_id} not found")

@router.post("/manipulate")
async def manipulate_document(
    request: DocumentManipulationRequest,
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Manipulasi dokumen berdasarkan pengaturan yang diberikan"""
    try:
        logger.info(f"Processing document manipulation request for file_id: {request.file_id}")
        
        # Get file path
        file_path = get_file_path_from_id(request.file_id)
        
        # Process document with enhanced service
        result = service.process_document_with_manipulation(
            input_file_path=file_path,
            settings=request.settings
        )
        
        if not result['success']:
            raise HTTPException(status_code=500, detail=result.get('error', 'Processing failed'))
        
        # If preview only, generate preview data
        if request.preview_only:
            preview_data = service.get_preview_data(result['output_path'])
            result['preview'] = preview_data
        
        return JSONResponse(content={
            "success": True,
            "message": "Document manipulated successfully",
            "data": result
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in document manipulation: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/split")
async def split_pdf_document(
    request: PDFSplitRequest,
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Split PDF dokumen menjadi beberapa file"""
    try:
        logger.info(f"Processing PDF split request for file_id: {request.file_id}")
        
        # Get file path
        file_path = get_file_path_from_id(request.file_id)
        
        # Prepare split settings
        split_settings = {
            'split_type': request.split_type,
            'split_value': request.split_value
        }
        
        if request.ranges:
            split_settings['ranges'] = request.ranges
        
        # Split PDF
        split_results = service.split_pdf(file_path, split_settings)
        
        if not split_results:
            raise HTTPException(status_code=500, detail="Failed to split PDF")
        
        return JSONResponse(content={
            "success": True,
            "message": f"PDF split into {len(split_results)} files",
            "data": {
                "original_file": request.file_id,
                "split_files": split_results,
                "total_files": len(split_results)
            }
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in PDF splitting: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/convert")
async def convert_document(
    request: DocumentConversionRequest,
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Konversi dokumen ke format yang diinginkan"""
    try:
        logger.info(f"Processing document conversion request for file_id: {request.file_id}")
        
        # Get file path
        file_path = get_file_path_from_id(request.file_id)
        
        # Prepare conversion settings
        conversion_settings = request.conversion_settings or {}
        conversion_settings['target_format'] = request.target_format
        
        # Process document (conversion is part of manipulation)
        result = service.process_document_with_manipulation(
            input_file_path=file_path,
            settings=conversion_settings
        )
        
        if not result['success']:
            raise HTTPException(status_code=500, detail=result.get('error', 'Conversion failed'))
        
        return JSONResponse(content={
            "success": True,
            "message": f"Document converted to {request.target_format} successfully",
            "data": result
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in document conversion: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/preview")
async def get_document_preview(
    request: PreviewRequest,
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Dapatkan preview dokumen dengan pengaturan tertentu"""
    try:
        logger.info(f"Generating preview for file_id: {request.file_id}, page: {request.page_number}")
        
        # Get file path
        file_path = get_file_path_from_id(request.file_id)
        
        # If settings provided, apply them first
        if request.settings:
            # Create temporary processed version
            temp_result = service.process_document_with_manipulation(
                input_file_path=file_path,
                settings=request.settings
            )
            
            if temp_result['success']:
                file_path = temp_result['output_path']
        
        # Generate preview
        preview_data = service.get_preview_data(
            pdf_path=file_path,
            page_num=request.page_number - 1  # Convert to 0-based
        )
        
        if not preview_data['success']:
            raise HTTPException(status_code=500, detail=preview_data.get('error', 'Preview generation failed'))
        
        return JSONResponse(content={
            "success": True,
            "message": "Preview generated successfully",
            "data": preview_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating preview: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/preview-processed")
async def get_processed_document_preview(
    request: DocumentManipulationRequest,
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Generate preview of processed document without saving permanently"""
    try:
        logger.info(f"Generating processed preview for file_id: {request.file_id}")
        
        # Get file path
        file_path = get_file_path_from_id(request.file_id)
        
        # Process document temporarily for preview
        result = service.process_document_with_manipulation(
            input_file_path=file_path,
            settings=request.settings
        )
        
        if not result['success']:
            raise HTTPException(status_code=500, detail=result.get('error', 'Processing failed'))
        
        # Generate preview data
        preview_data = service.get_preview_data(result['output_path'])
        
        # Create response with processed file info
        response_data = {
            "success": True,
            "message": "Processed preview generated successfully",
            "processed_file_id": result['output_filename'].replace('.pdf', ''),
            "preview_url": f"/api/files/{result['output_filename'].replace('.pdf', '')}/preview",
            "preview_data": preview_data,
            "processing_info": {
                "original_format": result.get('original_format'),
                "settings_applied": result.get('settings_applied'),
                "pages_count": result.get('pages_count'),
                "file_size": result.get('file_size')
            }
        }
        
        return JSONResponse(content=response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating processed preview: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/download/{file_id}")
async def download_processed_file(file_id: str):
    """Download file yang sudah diproses"""
    try:
        # Get file path
        file_path = get_file_path_from_id(file_id)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Get filename
        filename = Path(file_path).name
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/pdf'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading file: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/upload-and-process")
async def upload_and_process_document(
    file: UploadFile = File(...),
    settings: str = Form(...),
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Upload dan langsung proses dokumen"""
    try:
        # Parse settings
        try:
            settings_dict = json.loads(settings)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid settings JSON")
        
        # Save uploaded file temporarily
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_filename = f"{timestamp}_{file.filename}"
        temp_path = Path("temp") / temp_filename
        temp_path.parent.mkdir(exist_ok=True)
        
        # Write file
        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        logger.info(f"Uploaded file saved to: {temp_path}")
        
        # Process document
        result = service.process_document_with_manipulation(
            input_file_path=str(temp_path),
            settings=settings_dict
        )
        
        if not result['success']:
            # Cleanup temp file
            try:
                os.remove(temp_path)
            except:
                pass
            raise HTTPException(status_code=500, detail=result.get('error', 'Processing failed'))
        
        # Generate preview
        preview_data = service.get_preview_data(result['output_path'])
        result['preview'] = preview_data
        
        # Cleanup original temp file
        try:
            os.remove(temp_path)
        except:
            pass
        
        return JSONResponse(content={
            "success": True,
            "message": "File uploaded and processed successfully",
            "data": result
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in upload and process: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/supported-formats")
async def get_supported_formats():
    """Dapatkan daftar format yang didukung"""
    return JSONResponse(content={
        "success": True,
        "data": {
            "input_formats": {
                "pdf": [".pdf"],
                "word": [".docx", ".doc"],
                "excel": [".xlsx", ".xls", ".csv"],
                "image": [".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".gif"]
            },
            "output_formats": ["pdf"],
            "manipulation_features": [
                "color_conversion",
                "black_white_conversion",
                "grayscale_conversion",
                "brightness_adjustment",
                "contrast_adjustment",
                "page_splitting",
                "page_range_selection",
                "auto_rotation",
                "format_conversion"
            ],
            "paper_sizes": ["A4", "A3", "A5", "Letter", "Legal"],
            "orientations": ["portrait", "landscape"]
        }
    })

@router.delete("/cleanup")
async def cleanup_temp_files(
    older_than_hours: int = 24,
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Bersihkan file temporary yang lama"""
    try:
        cleaned_count = service.cleanup_temp_files(older_than_hours)
        
        return JSONResponse(content={
            "success": True,
            "message": f"Cleaned up {cleaned_count} temporary files",
            "data": {
                "cleaned_files": cleaned_count,
                "older_than_hours": older_than_hours
            }
        })
        
    except Exception as e:
        logger.error(f"Error cleaning up files: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/excel-preview")
async def get_excel_preview(
    request: ExcelPreviewRequest,
    service: EnhancedDocumentService = Depends(get_enhanced_document_service)
):
    """Get Excel file preview with sheet information"""
    try:
        import pandas as pd
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Get file path
        file_path = get_file_path_from_id(request.file_id)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Check if file is Excel
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File is not an Excel document")
        
        # Load Excel file
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        # Determine which sheet to preview
        if request.sheet_name:
            if request.sheet_name not in sheet_names:
                raise HTTPException(status_code=400, detail=f"Sheet '{request.sheet_name}' not found")
            active_sheet = workbook[request.sheet_name]
            sheet_index = sheet_names.index(request.sheet_name)
        elif request.sheet_index is not None:
            if request.sheet_index >= len(sheet_names) or request.sheet_index < 0:
                raise HTTPException(status_code=400, detail="Invalid sheet index")
            active_sheet = workbook[sheet_names[request.sheet_index]]
            sheet_index = request.sheet_index
        else:
            # Default to first sheet
            active_sheet = workbook.active
            sheet_index = 0
        
        # Get sheet data
        sheet_data = []
        max_row = min(active_sheet.max_row, 100)  # Limit to 100 rows for preview
        max_col = min(active_sheet.max_column, 20)  # Limit to 20 columns for preview
        
        for row in active_sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            sheet_data.append([str(cell) if cell is not None else "" for cell in row])
        
        # Get sheet info
        sheet_info = []
        for i, name in enumerate(sheet_names):
            sheet = workbook[name]
            sheet_info.append({
                "index": i,
                "name": name,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "is_active": i == sheet_index
            })
        
        workbook.close()
        
        return {
            "success": True,
            "file_id": request.file_id,
            "sheets": sheet_info,
            "active_sheet": {
                "name": sheet_names[sheet_index],
                "index": sheet_index,
                "data": sheet_data,
                "total_rows": active_sheet.max_row,
                "total_columns": active_sheet.max_column,
                "preview_rows": len(sheet_data),
                "preview_columns": len(sheet_data[0]) if sheet_data else 0
            },
            "file_info": {
                "name": os.path.basename(file_path),
                "size": os.path.getsize(file_path),
                "modified": datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat()
            }
        }
        
    except ImportError as e:
        logger.error(f"Required library not available: {e}")
        raise HTTPException(status_code=500, detail="Excel processing libraries not available")
    except Exception as e:
        logger.error(f"Error processing Excel preview: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process Excel preview: {str(e)}")

@router.post("/upload-spreadsheet")
async def upload_spreadsheet_for_preview(
    file: UploadFile = File(...),
    preserve_formatting: bool = Form(True),
    max_rows: int = Form(1000),
    max_columns: int = Form(50)
):
    """Upload Excel file and return formatted spreadsheet preview"""
    try:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        import uuid
        
        # Validate file type
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel document (.xlsx or .xls)")
        
        # Generate unique file ID
        file_id = f"spreadsheet_{uuid.uuid4().hex[:8]}"
        
        # Create temp directory if not exists
        temp_dir = Path("temp")
        temp_dir.mkdir(exist_ok=True)
        
        # Save uploaded file
        file_path = temp_dir / f"{file_id}_{file.filename}"
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Load Excel file with formatting
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        sheet_names = workbook.sheetnames
        
        # Process all sheets
        sheets_data = []
        for i, sheet_name in enumerate(sheet_names):
            sheet = workbook[sheet_name]
            
            # Get sheet dimensions
            max_row = min(sheet.max_row, max_rows) if sheet.max_row else 1
            max_col = min(sheet.max_column, max_columns) if sheet.max_column else 1
            
            # Extract data with formatting
            sheet_data = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    # Get cell value
                    value = cell.value if cell.value is not None else ""
                    
                    # Get cell formatting if preserve_formatting is True
                    cell_format = {}
                    if preserve_formatting:
                        # Font formatting
                        if cell.font:
                            cell_format['font'] = {
                                'bold': cell.font.bold,
                                'italic': cell.font.italic,
                                'underline': cell.font.underline,
                                'color': cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None,
                                'size': cell.font.size,
                                'name': cell.font.name
                            }
                        
                        # Fill/background color
                        if cell.fill and cell.fill.start_color:
                            cell_format['fill'] = {
                                'color': cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else None
                            }
                        
                        # Alignment
                        if cell.alignment:
                            cell_format['alignment'] = {
                                'horizontal': cell.alignment.horizontal,
                                'vertical': cell.alignment.vertical,
                                'wrap_text': cell.alignment.wrap_text
                            }
                        
                        # Border
                        if cell.border:
                            cell_format['border'] = {
                                'top': cell.border.top.style if cell.border.top else None,
                                'bottom': cell.border.bottom.style if cell.border.bottom else None,
                                'left': cell.border.left.style if cell.border.left else None,
                                'right': cell.border.right.style if cell.border.right else None
                            }
                    
                    row_data.append({
                        'value': str(value),
                        'format': cell_format if preserve_formatting else {},
                        'row': row_idx,
                        'col': col_idx
                    })
                
                sheet_data.append(row_data)
            
            sheets_data.append({
                'index': i,
                'name': sheet_name,
                'data': sheet_data,
                'total_rows': sheet.max_row or 0,
                'total_columns': sheet.max_column or 0,
                'preview_rows': len(sheet_data),
                'preview_columns': len(sheet_data[0]) if sheet_data else 0
            })
        
        workbook.close()
        
        # Store file path mapping
        file_mapping = {
            file_id: str(file_path)
        }
        
        return {
            "success": True,
            "file_id": file_id,
            "filename": file.filename,
            "sheets": sheets_data,
            "active_sheet": sheets_data[0] if sheets_data else None,
            "file_info": {
                "name": file.filename,
                "size": len(content),
                "sheets_count": len(sheets_data),
                "preserve_formatting": preserve_formatting,
                "max_rows": max_rows,
                "max_columns": max_columns
            },
            "message": "Spreadsheet uploaded and processed successfully"
        }
        
    except ImportError as e:
        logger.error(f"Required library not available: {e}")
        raise HTTPException(status_code=500, detail="Excel processing libraries not available")
    except Exception as e:
        logger.error(f"Error processing spreadsheet upload: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process spreadsheet: {str(e)}")

@router.post("/convert-excel-to-pdf")
async def convert_excel_to_pdf(request: ExcelToPDFRequest):
    """Konversi file Excel ke PDF dengan opsi preservasi format"""
    try:
        # Validasi file_id
        if not request.file_id:
            raise HTTPException(status_code=400, detail="file_id is required")
        
        # Cari file berdasarkan file_id
        file_path = None
        for temp_file in temp_files:
            if temp_file["id"] == request.file_id:
                file_path = temp_file["path"]
                break
        
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Validasi apakah file adalah Excel
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        # Import service konversi PDF
        from services.perfect_pdf_conversion_service import PerfectPDFConversionService
        
        # Inisialisasi service
        pdf_service = PerfectPDFConversionService()
        
        # Siapkan opsi konversi
        conversion_options = {
            'preserve_formatting': request.preserve_formatting,
            'preserve_charts': request.preserve_charts,
            'preserve_images': request.preserve_images,
            'preserve_formulas': request.preserve_formulas,
            'preserve_colors': request.preserve_colors,
            'preserve_fonts': request.preserve_fonts,
            'preserve_borders': request.preserve_borders,
            'preserve_alignment': request.preserve_alignment,
            'fit_to_page': request.fit_to_page,
            'high_quality': request.high_quality,
            'include_all_sheets': request.include_all_sheets,
            'page_orientation': request.page_orientation,
            'paper_size': request.paper_size,
            'margins': request.margins,
            'conversion_method': request.conversion_method
        }
        
        # Generate nama file output
        original_filename = os.path.basename(file_path)
        pdf_filename = os.path.splitext(original_filename)[0] + ".pdf"
        output_path = os.path.join(TEMP_DIR, f"converted_{uuid.uuid4().hex}_{pdf_filename}")
        
        # Lakukan konversi
        conversion_result = pdf_service.convert_excel_to_pdf_with_options(
            input_path=file_path,
            output_path=output_path,
            options=conversion_options
        )
        
        if conversion_result['status'] != 'success':
            raise HTTPException(
                status_code=500, 
                detail=f"Conversion failed: {conversion_result.get('error', 'Unknown error')}"
            )
        
        # Simpan file PDF ke temp_files
        pdf_file_id = str(uuid.uuid4())
        temp_files.append({
            "id": pdf_file_id,
            "filename": pdf_filename,
            "path": output_path,
            "upload_time": datetime.now(),
            "type": "application/pdf"
        })
        
        # Hitung ukuran file
        file_size = os.path.getsize(output_path)
        
        return {
            "status": "success",
            "message": "Excel file successfully converted to PDF",
            "pdf_file_id": pdf_file_id,
            "pdf_filename": pdf_filename,
            "file_size": file_size,
            "conversion_result": {
                "method": conversion_result.get('method', 'unknown'),
                "quality_score": conversion_result.get('quality_score', 0),
                "features_preserved": conversion_result.get('features_preserved', {}),
                "format_preservation": conversion_result.get('format_preservation', {})
            },
            "conversion_options": conversion_options,
            "download_url": f"/api/document-manipulation/download/{pdf_file_id}"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error converting Excel to PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@router.get("/health")
async def health_check():
    """Health check untuk document manipulation service"""
    try:
        # Test basic functionality
        temp_dir = Path("temp")
        temp_dir.mkdir(exist_ok=True)
        
        return JSONResponse(content={
            "success": True,
            "message": "Document manipulation service is healthy",
            "data": {
                "service_status": "active",
                "temp_directory": str(temp_dir),
                "temp_directory_exists": temp_dir.exists(),
                "timestamp": datetime.now().isoformat()
            }
        })
        
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))