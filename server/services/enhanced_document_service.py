import os
import logging
from typing import Dict, Any, Optional, Tuple, List, Union
from pathlib import Path
from datetime import datetime
import tempfile
import shutil
import io
import base64

# PDF processing libraries
import PyPDF2
import fitz  # PyMuPDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4, legal, A3, A5
from reportlab.lib.units import inch, cm, mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Spacer
from reportlab.lib import colors
from PIL import Image, ImageEnhance, ImageOps

# Document conversion libraries
from docx2pdf import convert as docx_to_pdf
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

logger = logging.getLogger(__name__)

class EnhancedDocumentService:
    """Enhanced service untuk manipulasi dan konversi dokumen dengan fitur lengkap"""
    
    def __init__(self, temp_dir: str = "temp"):
        self.temp_dir = Path(temp_dir)
        self.temp_dir.mkdir(exist_ok=True)
        
        # Supported formats
        self.supported_formats = {
            'pdf': ['.pdf'],
            'word': ['.docx', '.doc'],
            'excel': ['.xlsx', '.xls', '.csv'],
            'image': ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.gif']
        }
        
        # Paper sizes
        self.paper_sizes = {
            'A4': A4,
            'A3': A3, 
            'A5': A5,
            'Letter': letter,
            'Legal': legal
        }
        
        # Default settings
        self.default_settings = {
            'paper_size': 'A4',
            'orientation': 'portrait',
            'margin_top': 20,
            'margin_bottom': 20,
            'margin_left': 20,
            'margin_right': 20,
            'scale': 100,
            'color_mode': 'color',  # color, grayscale, black_white
            'quality': 'high',
            'copies': 1,
            'duplex': False,
            'fit_to_page': False,
            'center_horizontally': True,
            'center_vertically': True,
            # New manipulation settings
            'convert_to_bw': False,
            'split_pages': False,
            'page_range': 'all',  # 'all', '1-5', '1,3,5'
            'brightness': 0,  # -100 to 100
            'contrast': 0,    # -100 to 100
            'auto_rotate': False,
            'remove_blank_pages': False
        }
    
    def process_document_with_manipulation(self, input_file_path: str, settings: Dict[str, Any]) -> Dict[str, Any]:
        """Proses dokumen dengan manipulasi lengkap berdasarkan pengaturan"""
        try:
            input_path = Path(input_file_path)
            if not input_path.exists():
                raise FileNotFoundError(f"Input file not found: {input_file_path}")
            
            # Merge settings dengan default
            final_settings = {**self.default_settings, **settings}
            
            # Deteksi format file
            file_format = self._detect_file_format(input_path)
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{input_path.stem}_processed_{timestamp}.pdf"
            output_path = self.temp_dir / output_filename
            
            logger.info(f"Processing document: {input_file_path} -> {output_path}")
            logger.info(f"Detected format: {file_format}")
            logger.info(f"Settings: {final_settings}")
            
            # Step 1: Convert to PDF if needed
            pdf_path = self._convert_to_pdf(input_file_path, file_format, final_settings)
            
            # Step 2: Apply manipulations
            manipulated_pdf = self._apply_document_manipulations(pdf_path, final_settings)
            
            # Step 3: Apply print settings
            final_pdf = self._apply_print_settings(manipulated_pdf, str(output_path), final_settings)
            
            # Cleanup temporary files
            if pdf_path != input_file_path:
                try:
                    os.remove(pdf_path)
                except:
                    pass
            
            if manipulated_pdf != pdf_path and manipulated_pdf != str(output_path):
                try:
                    os.remove(manipulated_pdf)
                except:
                    pass
            
            # Get document info
            doc_info = self._get_document_info(str(output_path))
            
            return {
                'success': True,
                'input_path': input_file_path,
                'output_path': str(output_path),
                'output_filename': output_filename,
                'original_format': file_format,
                'settings_applied': final_settings,
                'file_size': os.path.getsize(output_path),
                'pages_count': doc_info.get('pages', 0),
                'created_at': datetime.now().isoformat(),
                'preview_available': True
            }
            
        except Exception as e:
            logger.error(f"Error processing document {input_file_path}: {e}")
            return {
                'success': False,
                'error': str(e),
                'input_path': input_file_path
            }
    
    def _detect_file_format(self, file_path: Path) -> str:
        """Deteksi format file berdasarkan ekstensi"""
        ext = file_path.suffix.lower()
        
        for format_type, extensions in self.supported_formats.items():
            if ext in extensions:
                return format_type
        
        return 'unknown'
    
    def _convert_to_pdf(self, input_path: str, file_format: str, settings: Dict[str, Any] = None) -> str:
        """Konversi berbagai format ke PDF"""
        if file_format == 'pdf':
            return input_path
        
        input_file = Path(input_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_output = self.temp_dir / f"{input_file.stem}_converted_{timestamp}.pdf"
        
        try:
            if file_format == 'word':
                self._convert_word_to_pdf(input_path, str(pdf_output), settings)
            elif file_format == 'excel':
                self._convert_excel_to_pdf(input_path, str(pdf_output))
            elif file_format == 'image':
                self._convert_image_to_pdf(input_path, str(pdf_output))
            else:
                raise ValueError(f"Unsupported format: {file_format}")
            
            logger.info(f"Converted {file_format} to PDF: {pdf_output}")
            return str(pdf_output)
            
        except Exception as e:
            logger.error(f"Error converting {file_format} to PDF: {e}")
            raise
    
    def _convert_word_to_pdf(self, input_path: str, output_path: str, settings: Dict[str, Any] = None):
        """Konversi Word document ke PDF dengan pengaturan cetak"""
        try:
            if settings and input_path.endswith('.docx'):
                # Apply Word-specific settings before conversion
                temp_docx_path = self._apply_word_settings(input_path, settings)
                docx_to_pdf(temp_docx_path, output_path)
                # Clean up temporary file
                if temp_docx_path != input_path:
                    os.remove(temp_docx_path)
            else:
                # Standard conversion without settings
                docx_to_pdf(input_path, output_path)
        except Exception as e:
            logger.error(f"Error converting Word to PDF: {e}")
            # Fallback: create a simple PDF with text content
            self._create_fallback_pdf(input_path, output_path, "Word Document")
    
    def _convert_excel_to_pdf(self, input_path: str, output_path: str):
        """Konversi Excel file ke PDF"""
        try:
            # Read Excel file
            if input_path.endswith('.csv'):
                df = pd.read_csv(input_path)
            else:
                df = pd.read_excel(input_path)
            
            # Create PDF using reportlab
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib import colors
            
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            elements = []
            
            # Convert DataFrame to list of lists
            data = [df.columns.tolist()] + df.values.tolist()
            
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
        except Exception as e:
            logger.error(f"Error converting Excel to PDF: {e}")
            self._create_fallback_pdf(input_path, output_path, "Excel Document")
    
    def _convert_image_to_pdf(self, input_path: str, output_path: str):
        """Konversi image ke PDF"""
        try:
            # Open and process image
            with Image.open(input_path) as img:
                # Convert to RGB if necessary
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Create PDF
                img.save(output_path, "PDF", resolution=100.0)
                
        except Exception as e:
            logger.error(f"Error converting image to PDF: {e}")
            raise
    
    def _apply_word_settings(self, input_path: str, settings: Dict[str, Any]) -> str:
        """Terapkan pengaturan cetak pada dokumen Word"""
        try:
            from docx import Document
            from docx.shared import Inches, Cm, Pt
            from docx.enum.section import WD_ORIENT
            from docx.enum.text import WD_ALIGN_PARAGRAPH
            
            # Load the document
            doc = Document(input_path)
            
            # Apply page settings
            for section in doc.sections:
                # Set orientation
                if settings.get('orientation') == 'landscape':
                    section.orientation = WD_ORIENT.LANDSCAPE
                    # Swap width and height for landscape
                    new_width, new_height = section.page_height, section.page_width
                    section.page_width = new_width
                    section.page_height = new_height
                else:
                    section.orientation = WD_ORIENT.PORTRAIT
                
                # Set page size based on paper_size
                paper_size = settings.get('paper_size', 'A4')
                if paper_size == 'A4':
                    section.page_width = Cm(21)
                    section.page_height = Cm(29.7)
                elif paper_size == 'LETTER':
                    section.page_width = Inches(8.5)
                    section.page_height = Inches(11)
                elif paper_size == 'LEGAL':
                    section.page_width = Inches(8.5)
                    section.page_height = Inches(14)
                elif paper_size == 'A3':
                    section.page_width = Cm(29.7)
                    section.page_height = Cm(42)
                elif paper_size == 'A5':
                    section.page_width = Cm(14.8)
                    section.page_height = Cm(21)
                
                # Set margins
                margin_top = settings.get('margin_top', 2.54)  # Default 1 inch in cm
                margin_bottom = settings.get('margin_bottom', 2.54)
                margin_left = settings.get('margin_left', 2.54)
                margin_right = settings.get('margin_right', 2.54)
                
                section.top_margin = Cm(margin_top)
                section.bottom_margin = Cm(margin_bottom)
                section.left_margin = Cm(margin_left)
                section.right_margin = Cm(margin_right)
            
            # Apply font and formatting settings if needed
            if settings.get('color_mode') == 'grayscale':
                # Convert colors to grayscale (simplified approach)
                for paragraph in doc.paragraphs:
                    for run in paragraph.runs:
                        if run.font.color and run.font.color.rgb:
                            # Convert to grayscale
                            rgb = run.font.color.rgb
                            gray = int(0.299 * rgb.red + 0.587 * rgb.green + 0.114 * rgb.blue)
                            run.font.color.rgb = (gray, gray, gray)
            
            # Save modified document to temporary file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            temp_path = self.temp_dir / f"modified_word_{timestamp}.docx"
            doc.save(str(temp_path))
            
            return str(temp_path)
            
        except Exception as e:
            logger.error(f"Error applying Word settings: {e}")
            # Return original path if modification fails
            return input_path
    
    def _create_fallback_pdf(self, input_path: str, output_path: str, doc_type: str):
        """Buat PDF fallback jika konversi gagal"""
        from reportlab.pdfgen import canvas
        
        c = canvas.Canvas(output_path, pagesize=A4)
        c.drawString(100, 750, f"Document: {Path(input_path).name}")
        c.drawString(100, 730, f"Type: {doc_type}")
        c.drawString(100, 710, "Original file could not be converted.")
        c.drawString(100, 690, "Please check the file format and try again.")
        c.save()
    
    def _apply_document_manipulations(self, pdf_path: str, settings: Dict[str, Any]) -> str:
        """Terapkan manipulasi dokumen (warna, split, dll)"""
        try:
            pdf_document = fitz.open(pdf_path)
            
            # Generate output path for manipulated PDF
            input_file = Path(pdf_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.temp_dir / f"{input_file.stem}_manipulated_{timestamp}.pdf"
            
            # Create new PDF document
            new_doc = fitz.open()
            
            # Get pages to process based on settings
            pages_to_process = self._get_pages_to_process(pdf_document, settings)
            
            for page_num in pages_to_process:
                if page_num < len(pdf_document):
                    page = pdf_document[page_num]
                    
                    # Apply manipulations to page
                    manipulated_page = self._manipulate_page(page, settings)
                    
                    # Add to new document
                    new_doc.insert_pdf(manipulated_page)
                    manipulated_page.close()
            
            # Save manipulated PDF
            new_doc.save(str(output_path))
            new_doc.close()
            pdf_document.close()
            
            logger.info(f"Applied manipulations: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error applying manipulations: {e}")
            return pdf_path  # Return original if manipulation fails
    
    def _manipulate_page(self, page: fitz.Page, settings: Dict[str, Any]) -> fitz.Document:
        """Manipulasi satu halaman PDF"""
        try:
            # Create new document for this page
            new_doc = fitz.open()
            new_page = new_doc.new_page(width=page.rect.width, height=page.rect.height)
            
            # Get page as pixmap for manipulation
            zoom = 2.0 if settings.get('quality') == 'high' else 1.5
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # Convert to PIL Image for advanced manipulation
            img_data = pix.tobytes("ppm")
            pil_img = Image.open(io.BytesIO(img_data))
            
            # Apply color manipulations
            color_mode = settings.get('color_mode')
            if settings.get('convert_to_bw') or color_mode == 'black_white' or (hasattr(color_mode, 'value') and color_mode.value == 'BLACK_WHITE'):
                pil_img = self._convert_to_black_white(pil_img)
            elif color_mode == 'grayscale' or (hasattr(color_mode, 'value') and color_mode.value == 'GRAYSCALE'):
                pil_img = pil_img.convert('L').convert('RGB')
            
            # Apply brightness and contrast
            if settings.get('brightness', 0) != 0:
                enhancer = ImageEnhance.Brightness(pil_img)
                factor = 1.0 + (settings['brightness'] / 100.0)
                pil_img = enhancer.enhance(factor)
            
            if settings.get('contrast', 0) != 0:
                enhancer = ImageEnhance.Contrast(pil_img)
                factor = 1.0 + (settings['contrast'] / 100.0)
                pil_img = enhancer.enhance(factor)
            
            # Auto rotate if needed
            if settings.get('auto_rotate'):
                pil_img = ImageOps.autocontrast(pil_img)
            
            # Convert back to pixmap
            img_bytes = io.BytesIO()
            pil_img.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            
            # Insert manipulated image into new page
            new_page.insert_image(page.rect, stream=img_bytes.getvalue())
            
            return new_doc
            
        except Exception as e:
            logger.error(f"Error manipulating page: {e}")
            # Return original page if manipulation fails
            fallback_doc = fitz.open()
            fallback_doc.insert_pdf(fitz.open(), from_page=page.number, to_page=page.number)
            return fallback_doc
    
    def _convert_to_black_white(self, pil_img: Image.Image, threshold: int = 128) -> Image.Image:
        """Konversi gambar ke hitam putih dengan threshold"""
        try:
            # Convert to grayscale first
            gray_img = pil_img.convert('L')
            
            # Apply threshold for black and white
            bw_img = gray_img.point(lambda x: 255 if x > threshold else 0, mode='1')
            
            # Convert back to RGB for consistency
            return bw_img.convert('RGB')
            
        except Exception as e:
            logger.error(f"Error converting to black and white: {e}")
            return pil_img.convert('L').convert('RGB')
    
    def _apply_print_settings(self, pdf_path: str, output_path: str, settings: Dict[str, Any]) -> str:
        """Terapkan pengaturan cetak (ukuran kertas, margin, dll)"""
        try:
            pdf_document = fitz.open(pdf_path)
            
            # Get target page size
            target_size = self._get_target_page_size(settings)
            
            # Create output document
            output_doc = fitz.open()
            
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                
                # Create new page with target size
                new_page = output_doc.new_page(width=target_size[0], height=target_size[1])
                
                # Calculate scaling and positioning
                scale_matrix = self._calculate_scale_matrix(page.rect, target_size, settings)
                
                # Calculate margins
                margin_left = settings.get('margin_left', 20) * mm
                margin_top = settings.get('margin_top', 20) * mm
                margin_right = settings.get('margin_right', 20) * mm
                margin_bottom = settings.get('margin_bottom', 20) * mm
                
                # Available space
                available_width = target_size[0] - margin_left - margin_right
                available_height = target_size[1] - margin_top - margin_bottom
                
                # Calculate scaled dimensions
                scaled_width = page.rect.width * scale_matrix.a
                scaled_height = page.rect.height * scale_matrix.d
                
                # When fit_to_page is enabled, center content to distribute margins evenly
                fit_to_page = settings.get('fit_to_page', False)
                
                if fit_to_page:
                    # For fit_to_page, always center to ensure even margins
                    x = margin_left + (available_width - scaled_width) / 2
                    y = margin_top + (available_height - scaled_height) / 2
                else:
                    # Use original centering settings
                    if settings.get('center_horizontally'):
                        x = margin_left + (available_width - scaled_width) / 2
                    else:
                        x = margin_left
                    
                    if settings.get('center_vertically'):
                        y = margin_top + (available_height - scaled_height) / 2
                    else:
                        y = margin_top
                
                # Create insert rect with correct position and scaled size
                insert_rect = fitz.Rect(x, y, x + scaled_width, y + scaled_height)
                
                # Render page at original size, then let insert_image handle the scaling
                pix = page.get_pixmap()
                
                new_page.insert_image(insert_rect, pixmap=pix)
            
            # Save final PDF
            output_doc.save(output_path)
            output_doc.close()
            pdf_document.close()
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error applying print settings: {e}")
            # Copy original file if print settings fail
            shutil.copy2(pdf_path, output_path)
            return output_path
    
    def _get_target_page_size(self, settings: Dict[str, Any]) -> Tuple[float, float]:
        """Dapatkan ukuran halaman target berdasarkan pengaturan"""
        paper_size = settings.get('paper_size', 'A4')
        orientation = settings.get('orientation', 'portrait')
        
        if paper_size in self.paper_sizes:
            width, height = self.paper_sizes[paper_size]
        else:
            width, height = A4  # Default
        
        if orientation == 'landscape':
            width, height = height, width
        
        return (width, height)
    
    def _calculate_scale_matrix(self, page_rect: fitz.Rect, target_size: Tuple[float, float], settings: Dict[str, Any]) -> fitz.Matrix:
        """Hitung matrix scaling untuk halaman"""
        scale = settings.get('scale', 100) / 100.0
        
        if settings.get('fit_to_page'):
            # Calculate available space after margins
            margin_left = settings.get('margin_left', 20) * mm
            margin_top = settings.get('margin_top', 20) * mm
            margin_right = settings.get('margin_right', 20) * mm
            margin_bottom = settings.get('margin_bottom', 20) * mm
            
            available_width = target_size[0] - margin_left - margin_right
            available_height = target_size[1] - margin_top - margin_bottom
            
            # Calculate scale based on available space, not full page size
            scale_x = available_width / page_rect.width
            scale_y = available_height / page_rect.height
            scale = min(scale_x, scale_y) * scale
        
        return fitz.Matrix(scale, scale)
    
    def _calculate_insert_rect(self, page_rect: fitz.Rect, target_size: Tuple[float, float], 
                              settings: Dict[str, Any], scale_matrix: fitz.Matrix) -> fitz.Rect:
        """Hitung posisi insert untuk halaman"""
        scaled_width = page_rect.width * scale_matrix.a
        scaled_height = page_rect.height * scale_matrix.d
        
        # Calculate margins
        margin_left = settings.get('margin_left', 20) * mm
        margin_top = settings.get('margin_top', 20) * mm
        margin_right = settings.get('margin_right', 20) * mm
        margin_bottom = settings.get('margin_bottom', 20) * mm
        
        # Available space
        available_width = target_size[0] - margin_left - margin_right
        available_height = target_size[1] - margin_top - margin_bottom
        
        # When fit_to_page is enabled, center content to distribute margins evenly
        fit_to_page = settings.get('fit_to_page', False)
        
        if fit_to_page:
            # For fit_to_page, always center to ensure even margins
            x = margin_left + (available_width - scaled_width) / 2
            y = margin_top + (available_height - scaled_height) / 2
        else:
            # Use original centering settings
            if settings.get('center_horizontally'):
                x = margin_left + (available_width - scaled_width) / 2
            else:
                x = margin_left
            
            if settings.get('center_vertically'):
                y = margin_top + (available_height - scaled_height) / 2
            else:
                y = margin_top
        
        return fitz.Rect(x, y, x + scaled_width, y + scaled_height)
    
    def _get_pages_to_process(self, pdf_document: fitz.Document, settings: Dict[str, Any]) -> List[int]:
        """Dapatkan daftar halaman yang akan diproses berdasarkan pengaturan"""
        total_pages = len(pdf_document)
        page_range_type = settings.get('page_range_type', 'all')
        page_range = settings.get('page_range', '')
        current_page = settings.get('current_page', 1)  # Default to page 1
        
        if page_range_type == 'all':
            return list(range(total_pages))
        elif page_range_type == 'current':
            # Convert to 0-based index and ensure it's within bounds
            current_idx = max(0, min(current_page - 1, total_pages - 1))
            return [current_idx]
        elif page_range_type == 'odd':
            return [i for i in range(total_pages) if (i + 1) % 2 == 1]  # 1, 3, 5, ...
        elif page_range_type == 'even':
            return [i for i in range(total_pages) if (i + 1) % 2 == 0]  # 2, 4, 6, ...
        elif page_range_type == 'custom' and page_range:
            return self._parse_page_range(page_range, total_pages)
        else:
            # Default ke semua halaman jika tidak valid
            return list(range(total_pages))
    
    def _parse_page_range(self, page_range: str, total_pages: int) -> List[int]:
        """Parse page range string seperti '1-5,8,11-13' menjadi list page numbers (0-indexed)"""
        pages = []
        
        try:
            # Split by comma untuk mendapatkan range atau single pages
            parts = [part.strip() for part in page_range.split(',')]
            
            for part in parts:
                if '-' in part:
                    # Range seperti '1-5'
                    start, end = part.split('-', 1)
                    start_page = max(1, min(int(start.strip()), total_pages))
                    end_page = max(1, min(int(end.strip()), total_pages))
                    
                    # Tambahkan range (convert to 0-indexed)
                    for page_num in range(start_page - 1, end_page):
                        if page_num not in pages:
                            pages.append(page_num)
                else:
                    # Single page seperti '8'
                    page_num = max(1, min(int(part.strip()), total_pages))
                    if (page_num - 1) not in pages:  # Convert to 0-indexed
                        pages.append(page_num - 1)
            
            # Sort pages
            pages.sort()
            return pages
            
        except (ValueError, IndexError) as e:
            logger.warning(f"Invalid page range format '{page_range}': {e}. Using all pages.")
            return list(range(total_pages))
    
    def _get_document_info(self, pdf_path: str) -> Dict[str, Any]:
        """Dapatkan informasi dokumen PDF"""
        try:
            pdf_document = fitz.open(pdf_path)
            info = {
                'pages': len(pdf_document),
                'title': pdf_document.metadata.get('title', ''),
                'author': pdf_document.metadata.get('author', ''),
                'subject': pdf_document.metadata.get('subject', ''),
                'creator': pdf_document.metadata.get('creator', ''),
                'producer': pdf_document.metadata.get('producer', ''),
                'creation_date': pdf_document.metadata.get('creationDate', ''),
                'modification_date': pdf_document.metadata.get('modDate', '')
            }
            pdf_document.close()
            return info
        except Exception as e:
            logger.error(f"Error getting document info: {e}")
            return {'pages': 0}
    
    def split_pdf(self, pdf_path: str, split_settings: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Split PDF menjadi beberapa file berdasarkan pengaturan"""
        try:
            pdf_document = fitz.open(pdf_path)
            total_pages = len(pdf_document)
            
            split_type = split_settings.get('split_type', 'pages')  # 'pages', 'range', 'size'
            split_value = split_settings.get('split_value', 1)
            
            results = []
            input_file = Path(pdf_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if split_type == 'pages':
                # Split every N pages
                pages_per_file = int(split_value)
                for i in range(0, total_pages, pages_per_file):
                    end_page = min(i + pages_per_file - 1, total_pages - 1)
                    output_filename = f"{input_file.stem}_part_{i//pages_per_file + 1}_{timestamp}.pdf"
                    output_path = self.temp_dir / output_filename
                    
                    # Create new document with selected pages
                    new_doc = fitz.open()
                    new_doc.insert_pdf(pdf_document, from_page=i, to_page=end_page)
                    new_doc.save(str(output_path))
                    new_doc.close()
                    
                    results.append({
                        'filename': output_filename,
                        'path': str(output_path),
                        'pages': f"{i+1}-{end_page+1}",
                        'page_count': end_page - i + 1
                    })
            
            elif split_type == 'range':
                # Split by specific page ranges
                ranges = split_settings.get('ranges', [])
                for idx, page_range in enumerate(ranges):
                    start, end = map(int, page_range.split('-'))
                    start -= 1  # Convert to 0-based
                    end -= 1
                    
                    if 0 <= start < total_pages and 0 <= end < total_pages:
                        output_filename = f"{input_file.stem}_range_{start+1}-{end+1}_{timestamp}.pdf"
                        output_path = self.temp_dir / output_filename
                        
                        new_doc = fitz.open()
                        new_doc.insert_pdf(pdf_document, from_page=start, to_page=end)
                        new_doc.save(str(output_path))
                        new_doc.close()
                        
                        results.append({
                            'filename': output_filename,
                            'path': str(output_path),
                            'pages': f"{start+1}-{end+1}",
                            'page_count': end - start + 1
                        })
            
            pdf_document.close()
            
            logger.info(f"Split PDF into {len(results)} files")
            return results
            
        except Exception as e:
            logger.error(f"Error splitting PDF: {e}")
            return []
    
    def get_preview_data(self, pdf_path: str, page_num: int = 0) -> Dict[str, Any]:
        """Dapatkan data preview untuk halaman PDF"""
        try:
            pdf_document = fitz.open(pdf_path)
            
            if page_num >= len(pdf_document):
                page_num = 0
            
            page = pdf_document[page_num]
            
            # Render page as image
            zoom = 1.5  # Good quality for preview
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # Convert to base64 for web display
            img_data = pix.tobytes("png")
            img_base64 = base64.b64encode(img_data).decode('utf-8')
            
            pdf_document.close()
            
            return {
                'success': True,
                'page_number': page_num + 1,
                'total_pages': len(pdf_document),
                'image_data': f"data:image/png;base64,{img_base64}",
                'page_width': pix.width,
                'page_height': pix.height
            }
            
        except Exception as e:
            logger.error(f"Error generating preview: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_excel_preview(self, file_path: str, sheet_name: str = None, sheet_index: int = None) -> Dict[str, Any]:
        """Generate preview untuk Excel file"""
        try:
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            
            # Determine which sheet to use
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    return {
                        'success': False,
                        'error': f'Sheet "{sheet_name}" not found'
                    }
            elif sheet_index is not None:
                if 0 <= sheet_index < len(workbook.sheetnames):
                    sheet = workbook.worksheets[sheet_index]
                else:
                    return {
                        'success': False,
                        'error': f'Sheet index {sheet_index} out of range'
                    }
            else:
                sheet = workbook.active
            
            # Extract data (maksimal 20 baris untuk preview)
            preview_data = []
            max_rows = min(20, sheet.max_row)
            max_cols = min(15, sheet.max_column)
            
            for row in range(1, max_rows + 1):
                row_data = []
                for col in range(1, max_cols + 1):
                    cell = sheet.cell(row=row, column=col)
                    row_data.append(str(cell.value) if cell.value is not None else "")
                preview_data.append(row_data)
            
            return {
                'success': True,
                'type': 'excel',
                'preview_available': True,
                'sheet_names': workbook.sheetnames,
                'active_sheet': sheet.title,
                'preview_data': preview_data,
                'total_rows': sheet.max_row,
                'total_columns': sheet.max_column,
                'displayed_rows': max_rows,
                'displayed_columns': max_cols
            }
            
        except Exception as e:
            logger.error(f"Error generating Excel preview: {e}")
            return {
                'success': False,
                'type': 'error',
                'preview_available': False,
                'error': str(e)
            }
    
    def cleanup_temp_files(self, older_than_hours: int = 24) -> int:
        """Bersihkan file temporary yang lama"""
        try:
            current_time = datetime.now()
            cleaned_count = 0
            
            for file_path in self.temp_dir.glob('*'):
                if file_path.is_file():
                    file_age = current_time - datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_age.total_seconds() > (older_than_hours * 3600):
                        try:
                            file_path.unlink()
                            cleaned_count += 1
                        except Exception as e:
                            logger.warning(f"Could not delete {file_path}: {e}")
            
            logger.info(f"Cleaned up {cleaned_count} temporary files")
            return cleaned_count
            
        except Exception as e:
            logger.error(f"Error cleaning up temp files: {e}")
            return 0
    
    def __del__(self):
        """Cleanup saat object dihapus"""
        try:
            self.cleanup_temp_files(0)  # Clean all temp files
        except:
            pass