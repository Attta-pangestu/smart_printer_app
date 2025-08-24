import openpyxl
import json
import os
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
import tempfile

logger = logging.getLogger(__name__)

class ExcelVisualService:
    """Service untuk preview Excel dan pemilihan area print visual"""
    
    def __init__(self, temp_dir: str = None):
        self.temp_dir = Path(temp_dir) if temp_dir else Path(tempfile.gettempdir())
        self.temp_dir.mkdir(exist_ok=True)
        
    def read_excel_data(self, file_path: str, sheet_name: str = None, max_rows: int = 100, max_cols: int = 50) -> Dict[str, Any]:
        """Membaca data Excel untuk ditampilkan di frontend"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Jika sheet_name tidak diberikan, gunakan sheet pertama
            if sheet_name is None:
                sheet_name = workbook.sheetnames[0]
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan")
            
            worksheet = workbook[sheet_name]
            
            # Mendapatkan dimensi aktual worksheet
            max_row = min(worksheet.max_row, max_rows)
            max_col = min(worksheet.max_column, max_cols)
            
            # Membaca data sel
            data = []
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    # Konversi nilai ke string yang bisa di-serialize JSON
                    if cell_value is None:
                        cell_value = ""
                    elif isinstance(cell_value, (int, float)):
                        cell_value = str(cell_value)
                    elif isinstance(cell_value, datetime):
                        cell_value = cell_value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        cell_value = str(cell_value)
                    
                    # Informasi sel dengan koordinat
                    cell_info = {
                        'value': cell_value,
                        'address': f"{self._get_column_letter(col)}{row}",
                        'row': row,
                        'col': col,
                        'has_value': bool(cell.value is not None and str(cell.value).strip())
                    }
                    
                    row_data.append(cell_info)
                data.append(row_data)
            
            # Informasi sheet
            sheet_info = {
                'name': sheet_name,
                'max_row': worksheet.max_row,
                'max_col': worksheet.max_column,
                'displayed_rows': max_row,
                'displayed_cols': max_col,
                'data': data,
                'column_headers': [self._get_column_letter(i) for i in range(1, max_col + 1)],
                'row_headers': list(range(1, max_row + 1))
            }
            
            workbook.close()
            
            return {
                'success': True,
                'sheet_info': sheet_info,
                'available_sheets': workbook.sheetnames
            }
            
        except Exception as e:
            logger.error(f"Error reading Excel data: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_excel_sheets_info(self, file_path: str) -> Dict[str, Any]:
        """Mendapatkan informasi semua sheet dalam file Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            sheets_info = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_info = {
                    'name': sheet_name,
                    'max_row': worksheet.max_row,
                    'max_col': worksheet.max_column,
                    'has_data': worksheet.max_row > 1 or worksheet.max_column > 1
                }
                sheets_info.append(sheet_info)
            
            workbook.close()
            
            return {
                'success': True,
                'sheets': sheets_info,
                'total_sheets': len(sheets_info)
            }
            
        except Exception as e:
            logger.error(f"Error getting Excel sheets info: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def validate_cell_range(self, file_path: str, sheet_name: str, cell_range: str) -> Dict[str, Any]:
        """Validasi range sel yang dipilih"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan")
            
            worksheet = workbook[sheet_name]
            
            # Parse range sel (contoh: A1:C10)
            if ':' in cell_range:
                start_cell, end_cell = cell_range.split(':')
                start_row, start_col = self._parse_cell_address(start_cell)
                end_row, end_col = self._parse_cell_address(end_cell)
            else:
                # Single cell
                start_row, start_col = self._parse_cell_address(cell_range)
                end_row, end_col = start_row, start_col
            
            # Validasi range
            if start_row > worksheet.max_row or start_col > worksheet.max_column:
                raise ValueError("Range sel melebihi data yang tersedia")
            
            # Hitung jumlah sel yang akan diproses
            total_cells = (end_row - start_row + 1) * (end_col - start_col + 1)
            
            workbook.close()
            
            return {
                'success': True,
                'valid': True,
                'start_row': start_row,
                'start_col': start_col,
                'end_row': end_row,
                'end_col': end_col,
                'total_cells': total_cells,
                'range_description': f"{cell_range} ({total_cells} sel)"
            }
            
        except Exception as e:
            logger.error(f"Error validating cell range: {e}")
            return {
                'success': False,
                'valid': False,
                'error': str(e)
            }
    
    def _get_column_letter(self, col_num: int) -> str:
        """Konversi nomor kolom ke huruf (1=A, 2=B, dst.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def _parse_cell_address(self, cell_address: str) -> Tuple[int, int]:
        """Parse alamat sel (contoh: A1) menjadi (row, col)"""
        col_str = ""
        row_str = ""
        
        for char in cell_address:
            if char.isalpha():
                col_str += char
            elif char.isdigit():
                row_str += char
        
        # Konversi kolom huruf ke nomor
        col_num = 0
        for char in col_str:
            col_num = col_num * 26 + (ord(char.upper()) - ord('A') + 1)
        
        row_num = int(row_str)
        
        return row_num, col_num
    
    def get_cell_range_data(self, file_path: str, sheet_name: str, cell_range: str) -> Dict[str, Any]:
        """Mendapatkan data dari range sel tertentu"""
        try:
            validation = self.validate_cell_range(file_path, sheet_name, cell_range)
            if not validation['success'] or not validation['valid']:
                return validation
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook[sheet_name]
            
            start_row = validation['start_row']
            start_col = validation['start_col']
            end_row = validation['end_row']
            end_col = validation['end_col']
            
            # Membaca data dari range
            range_data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    if cell_value is None:
                        cell_value = ""
                    elif isinstance(cell_value, (int, float)):
                        cell_value = str(cell_value)
                    elif isinstance(cell_value, datetime):
                        cell_value = cell_value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        cell_value = str(cell_value)
                    
                    row_data.append({
                        'value': cell_value,
                        'address': f"{self._get_column_letter(col)}{row}",
                        'row': row,
                        'col': col
                    })
                range_data.append(row_data)
            
            workbook.close()
            
            return {
                'success': True,
                'range_data': range_data,
                'range_info': {
                    'start_row': start_row,
                    'start_col': start_col,
                    'end_row': end_row,
                    'end_col': end_col,
                    'total_rows': end_row - start_row + 1,
                    'total_cols': end_col - start_col + 1,
                    'cell_range': cell_range
                }
            }
            
        except Exception as e:
            logger.error(f"Error getting cell range data: {e}")
            return {
                'success': False,
                'error': str(e)
            }